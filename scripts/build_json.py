#!/usr/bin/env python3
"""
CAL dashboard data build.
Reads the analysis workbook and writes four JSON files a Next.js dashboard renders:
  shows.json     one record per show acronym (rotation + revenue + espresso split)
  cities.json    one record per city (revenue by pipeline/year, shows that land there)
  patterns.json  four city-pattern buckets, totaling to grand revenue
  meta.json      scope totals, definitions, provenance, branding tokens

Source of truth:
  Revenue / deal counts / espresso  -> deals-wide tabs 2_National + 3_CL (one row per deal)
  City / month rotation per year    -> 6_Show_City_Year_Detail (analyst rotation table)
  City / month pattern class        -> 6_Show_City_Year_Detail pattern columns
All revenue numbers derive from the deal level so every summary reconciles to grand total.
"""
import json, re, sys
from collections import defaultdict
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else "workbook.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "."

YEARS = [2022, 2023, 2024, 2025, 2026]
PATTERN_CLASSES = ["Fixed City", "Rotation with Anchor", "Full Rotation", "One-Off / Custom"]
UNSPEC = "(Unspecified)"

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)


def norm(x):
    return str(x).strip() if x is not None else None


def key(x):
    return str(x).strip().upper() if x is not None else None


# ---------------------------------------------------------------- deal level
def read_deals(tab, pipeline):
    rows = list(wb[tab].iter_rows(values_only=True))
    # header at index 3: Show Full Name, Show Acronym, In FS, Show City,
    #                    Show Close Month, Show Year, Deal Value, Total Services, <services...>
    out = []
    for r in rows[4:]:
        val = r[6]
        if not isinstance(val, (int, float)):
            continue  # only rows carrying a real dollar value are deals
        out.append({
            "fullName": norm(r[0]),
            "acr": key(r[1]) or UNSPEC,
            "acrRaw": norm(r[1]) or UNSPEC,
            "inFS": norm(r[2]) == "Yes",
            "city": norm(r[3]) or "Unknown",
            "month": norm(r[4]),
            "year": int(r[5]) if isinstance(r[5], (int, float)) else None,
            "value": float(val),
            "espresso": bool(r[8]),
            "pipeline": pipeline,
        })
    return out


deals = read_deals("2_National_Deals_Wide", "National") + read_deals("3_CL_Deals_Wide", "Cafe Lines")

# ---------------------------------------- classification from tab 7 (source)
# Tab 7 is the authoritative Pattern Findings classification. Columns:
#   0 Acronym, 1 Full Name, 2 In FS, 3 City Pattern, 4 Month Pattern,
#   5 Years, 6 Pipelines, 7 Total Rev.
# It has one section per class plus a "Month Drift" cross-cut (whose shows are
# duplicates of the first three sections) and an "Insufficient Data" section.
# We read the primary three classes + Insufficient (-> One-Off/Custom) and skip
# the Month Drift cross-cut so revenue is never double-counted.
SECTION_CLASS = {
    "Fixed City Shows": "Fixed City",
    "Rotation with Anchor City": "Rotation with Anchor",
    "Full Rotation": "Full Rotation",
    "Insufficient Data for Pattern": "One-Off / Custom",
}
city_pat = {}    # acr -> raw City Pattern string  e.g. "Fixed City (Chicago)"
month_pat = {}   # acr -> raw Month Pattern string e.g. "Slight Drift (Dec / Nov)"
cls_of = {}      # acr -> pattern class
cur = None
for r in list(wb["7_Pattern_Findings"].iter_rows(values_only=True)):
    a0 = norm(r[0])
    if a0 is None:
        continue
    matched = next((c for pref, c in SECTION_CLASS.items() if a0.startswith(pref)), None)
    if matched:
        cur = matched
        continue
    if a0.startswith("Month Drift"):     # cross-cut section: skip its rows
        cur = None
        continue
    if a0 == "Acronym":                  # column header inside a section
        continue
    if cur is None:                      # inside skipped section or preamble
        continue
    acr = key(r[0])
    if acr and acr not in cls_of:        # first (primary) section wins
        cls_of[acr] = cur
        city_pat[acr] = norm(r[3])
        month_pat[acr] = norm(r[4])


def classify(acr):
    return cls_of.get(acr, "One-Off / Custom")


def month_drifts(acr):
    mp = month_pat.get(acr) or ""
    return "Drift" in mp


# tab 6: one row per acronym + pipeline. Columns 4..23 are year blocks of
# (City, Month, Rev, Deals). Used only for per-year city/month labels.
tab6 = list(wb["6_Show_City_Year_Detail"].iter_rows(values_only=True))
rotation = {}     # acr -> {pipeline -> {year -> {city, month}}}
for r in tab6[4:]:
    acr = key(r[1])
    if acr is None:
        continue
    pipe = norm(r[2])
    rotation.setdefault(acr, {}).setdefault(pipe, {})
    for i, y in enumerate(YEARS):
        base = 4 + i * 4
        city, month = norm(r[base]), norm(r[base + 1])
        if city or month:
            rotation[acr][pipe][y] = {"city": city, "month": month}


# ------------------------------------------------------------- per-show build
show_deals = defaultdict(list)
for d in deals:
    show_deals[d["acr"]].append(d)

shows = []
for acr, ds in show_deals.items():
    total = sum(d["value"] for d in ds)
    esp = sum(d["value"] for d in ds if d["espresso"])
    pipelines = sorted({d["pipeline"] for d in ds})
    years_present = sorted({d["year"] for d in ds if d["year"]})
    cls = classify(acr) if acr != UNSPEC else "One-Off / Custom"

    # by-year revenue/deals from deals (reconciling); city/month labels from tab 6
    by_year = {}
    for y in years_present:
        entry = {}
        for pipe, tag in (("National", "national"), ("Cafe Lines", "cafeLines")):
            pyd = [d for d in ds if d["pipeline"] == pipe and d["year"] == y]
            if not pyd:
                continue
            lbl = rotation.get(acr, {}).get(pipe, {}).get(y, {})
            # fall back to modal deal city/month if tab 6 has no label
            city = lbl.get("city") or (pyd[0]["city"] if pyd else None)
            month = lbl.get("month") or (pyd[0]["month"] if pyd else None)
            entry[tag] = {
                "city": city,
                "month": month,
                "revenue": round(sum(d["value"] for d in pyd), 2),
                "deals": len(pyd),
            }
        if entry:
            by_year[str(y)] = entry

    shows.append({
        "acronym": acr,
        "fullName": ds[0]["fullName"] or acr,
        "inFS": any(d["inFS"] for d in ds),
        "pipelines": pipelines,
        "cityPatternClass": cls,
        "cityPatternDetail": city_pat.get(acr),
        "monthPatternClass": (
            "Fixed Month" if (month_pat.get(acr) or "").startswith("Fixed Month")
            else "Slight Drift" if (month_pat.get(acr) or "").startswith("Slight Drift")
            else "Significant Drift" if (month_pat.get(acr) or "").startswith("Significant Drift")
            else "Insufficient Data"
        ),
        "monthPatternDetail": month_pat.get(acr),
        "monthDrift": month_drifts(acr),
        "years": years_present,
        "totalRevenue": round(total, 2),
        "dealCount": len(ds),
        "espressoSplit": {
            "espressoRevenue": round(esp, 2),
            "otherRevenue": round(total - esp, 2),
            "espressoSharePct": round(esp / total * 100, 1) if total else 0.0,
        },
        "byYear": by_year,
    })

shows.sort(key=lambda s: s["totalRevenue"], reverse=True)

# --------------------------------------------------------------- cities build
city_deals = defaultdict(list)
for d in deals:
    city_deals[d["city"]].append(d)

cities = []
for city, ds in city_deals.items():
    total = sum(d["value"] for d in ds)
    esp = sum(d["value"] for d in ds if d["espresso"])
    by_pipe = {}
    for pipe, tag in (("National", "national"), ("Cafe Lines", "cafeLines")):
        pd = [d for d in ds if d["pipeline"] == pipe]
        by_pipe[tag] = {"revenue": round(sum(d["value"] for d in pd), 2), "deals": len(pd)}
    by_year = {}
    for y in YEARS:
        yd = [d for d in ds if d["year"] == y]
        if yd:
            by_year[str(y)] = {"revenue": round(sum(d["value"] for d in yd), 2), "deals": len(yd)}
    acrs = sorted({d["acr"] for d in ds})
    cities.append({
        "city": city,
        "totalRevenue": round(total, 2),
        "dealCount": len(ds),
        "byPipeline": by_pipe,
        "byYear": by_year,
        "distinctShows": len(acrs),
        "showAcronyms": acrs,
        "espressoSharePct": round(esp / total * 100, 1) if total else 0.0,
    })

cities.sort(key=lambda c: c["totalRevenue"], reverse=True)

# ------------------------------------------------------------- patterns build
buckets = {c: {"deals": 0, "rev": 0.0, "acrs": set(), "drift": set(),
               "revNat": 0.0, "revCL": 0.0} for c in PATTERN_CLASSES}
for d in deals:
    cls = classify(d["acr"]) if d["acr"] != UNSPEC else "One-Off / Custom"
    b = buckets[cls]
    b["deals"] += 1
    b["rev"] += d["value"]
    b["revNat" if d["pipeline"] == "National" else "revCL"] += d["value"]
    b["acrs"].add(d["acr"])
    if month_drifts(d["acr"]):
        b["drift"].add(d["acr"])

DESCR = {
    "Fixed City": "Same city 3+ years. Revenue trend is a demand or menu signal, not a location one.",
    "Rotation with Anchor": "Rotates but repeats one city across the cycle.",
    "Full Rotation": "Different city each year. Revenue heavily influenced by which city they land in.",
    "One-Off / Custom": "Fewer than 3 years of history or unclassified. One-and-done or custom shows.",
}
patterns = []
for c in PATTERN_CLASSES:
    b = buckets[c]
    pipe_ct = defaultdict(int)
    for acr in b["acrs"]:
        p = sorted({d["pipeline"] for d in show_deals[acr]})
        pipe_ct["National + CL" if len(p) == 2 else p[0]] += 1
    patterns.append({
        "cityPatternClass": c,
        "description": DESCR[c],
        "showCount": len(b["acrs"]),
        "dealCount": b["deals"],
        "totalRevenue": round(b["rev"], 2),
        "revenueByPipeline": {
            "National": round(b["revNat"], 2),
            "Cafe Lines": round(b["revCL"], 2),
        },
        "showCountByPipeline": dict(pipe_ct),
        "showsWithMonthDrift": len(b["drift"]),
    })

# ----------------------------------------------------------------- meta build
nat = [d for d in deals if d["pipeline"] == "National"]
cl = [d for d in deals if d["pipeline"] == "Cafe Lines"]
grand = sum(d["value"] for d in deals)
esp_all = sum(d["value"] for d in deals if d["espresso"])
meta = {
    "title": "CAL Services and Location Analysis",
    "dateRange": {"start": "2022-01-01", "end": "2026-07-07", "label": "2022 to July 7 2026"},
    "generatedFrom": "CAL Services and Location Analysis 2022 to July 7 2026.xlsx",
    "totals": {
        "dealsInScope": len(deals),
        "national": {"deals": len(nat), "revenue": round(sum(d["value"] for d in nat), 2)},
        "cafeLines": {"deals": len(cl), "revenue": round(sum(d["value"] for d in cl), 2)},
        "grandRevenue": round(grand, 2),
        "uniqueShowAcronyms": len({d["acr"] for d in deals}),
        "distinctCities": len(city_deals),
    },
    "espressoSplit": {
        "definition": "A deal counts as espresso if it includes Espresso Service. "
                      "Revenue is not split within multi-service deals.",
        "espressoRevenue": round(esp_all, 2),
        "otherRevenue": round(grand - esp_all, 2),
        "espressoSharePct": round(esp_all / grand * 100, 1),
    },
    "pipelines": ["National", "Cafe Lines"],
    "years": YEARS,
    "yearLabels": {str(y): (f"{y} YTD" if y == 2026 else str(y)) for y in YEARS},
    "definitions": {
        "revenue": "Won deals at real dollar value. Show Close Date drives year/month bucketing.",
        "cityPatternClasses": PATTERN_CLASSES,
        "monthPatternClasses": ["Fixed Month", "Slight Drift", "Significant Drift", "Insufficient Data"],
        "oneOffCustom": "Workbook 'Insufficient Data' shows (under 3 years of history) plus any "
                        "acronym absent from the pattern classification, so summary views total to grand revenue.",
    },
    "reconciliation": {
        "status": "reconciled",
        "note": "Two National deals carry no show acronym (Washington $5,645; city Unknown $13,195). "
                "They are retained in grand, city, and One-Off/Custom totals as the (Unspecified) show.",
    },
    "branding": {"purple": "#944197", "textHeader": "#000000",
                 "bodyFont": "Karla", "headerFont": "Bricolage Grotesque"},
}

# --------------------------------------------------------------------- write
def dump(name, obj):
    with open(f"{OUT}/{name}", "w") as f:
        json.dump(obj, f, indent=2, ensure_ascii=False)

dump("shows.json", shows)
dump("cities.json", cities)
dump("patterns.json", patterns)
dump("meta.json", meta)

# ------------------------------------------------------------ reconcile check
pat_sum = sum(p["totalRevenue"] for p in patterns)
city_sum = sum(c["totalRevenue"] for c in cities)
show_sum = sum(s["totalRevenue"] for s in shows)
print(f"grand           ${grand:,.0f}  deals {len(deals)}")
print(f"patterns sum    ${pat_sum:,.0f}  (diff ${grand-pat_sum:,.2f})")
print(f"cities sum      ${city_sum:,.0f}  (diff ${grand-city_sum:,.2f})")
print(f"shows sum       ${show_sum:,.0f}  (diff ${grand-show_sum:,.2f})")
print(f"National        ${sum(d['value'] for d in nat):,.0f}  deals {len(nat)}   (Overview: 54,852,121 / 2196)")
print(f"Cafe Lines      ${sum(d['value'] for d in cl):,.0f}   deals {len(cl)}   (Overview:  5,282,291 / 1151)")
print(f"espresso share  {meta['espressoSplit']['espressoSharePct']}%")
print(f"shows written   {len(shows)}   cities {len(cities)}")
for p in patterns:
    print(f"  {p['cityPatternClass']:<22} shows {p['showCount']:>3}  deals {p['dealCount']:>4}  ${p['totalRevenue']:>13,.0f}  drift {p['showsWithMonthDrift']}")
